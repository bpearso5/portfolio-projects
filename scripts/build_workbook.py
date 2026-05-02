"""Build project_tracker_dashboard.xlsx from the synthetic CSVs.

Tabs:
  README                — what this workbook is and how to use it
  Data_Projects         — cleaned projects with formula columns (pct_complete, pct_burned, health_flag)
  Data_Tasks            — tasks
  Data_TimeEntries_Wk   — time aggregated per employee per week with capacity flag
  Data_Budget           — monthly budget vs. actual
  Dashboard             — KPI cards, summary tables, charts
  Exec_OnePager         — print-ready summary

Run from anywhere:
    python3 scripts/build_workbook.py
"""

import csv
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).parent.parent          # ..  (script lives in scripts/)
DATA_DIR = ROOT / "data" / "raw"
OUT_PATH = ROOT / "project_tracker_dashboard.xlsx"

# ---------- helpers ----------
def read_csv(path):
    with open(path) as f:
        return list(csv.DictReader(f))

def parse_date(s):
    if not s:
        return None
    return datetime.strptime(s, "%Y-%m-%d").date()

def clean_status(s):
    s = (s or "").strip().upper()
    if s in ("IN PROGRESS","IN-PROGRESS","INPROGRESS"): return "In Progress"
    if s in ("NOT STARTED","NEW"): return "Not Started"
    if s in ("ON HOLD","PAUSED"): return "On Hold"
    if s in ("COMPLETED","DONE","CLOSED"): return "Completed"
    if s in ("CANCELLED","CANCELED"): return "Cancelled"
    return "Unknown"

# ---------- read sources ----------
employees = read_csv(DATA_DIR / "employees.csv")
projects = read_csv(DATA_DIR / "projects.csv")
tasks = read_csv(DATA_DIR / "tasks.csv")
time_entries = read_csv(DATA_DIR / "time_entries.csv")
budget = read_csv(DATA_DIR / "project_budget_actuals.csv")

emp_by_id = {int(e["employee_id"]): e for e in employees}

# pre-compute pct_complete and pct_burned per project (so we can demo formulas later)
task_by_proj = defaultdict(lambda: {"total":0,"done":0})
for t in tasks:
    pid = int(t["project_id"])
    task_by_proj[pid]["total"] += 1
    if t["status"] == "Done":
        task_by_proj[pid]["done"] += 1

budget_by_proj = defaultdict(lambda: {"plan":0.0,"actual":0.0})
for b in budget:
    pid = int(b["project_id"])
    budget_by_proj[pid]["plan"]   += float(b["budgeted_amount"])
    budget_by_proj[pid]["actual"] += float(b["actual_amount"])

# weekly hours per employee
hours_wk = defaultdict(float)
for te in time_entries:
    eid = int(te["employee_id"])
    d = parse_date(te["entry_date"])
    week_start = d - timedelta(days=d.weekday())
    hours_wk[(eid, week_start)] += float(te["hours_logged"])

# ---------- build workbook ----------
wb = Workbook()

# ----- styles -----
BRAND = "1F4E78"
HDR_FILL = PatternFill("solid", fgColor=BRAND)
HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial")
TITLE_FONT = Font(bold=True, size=18, color=BRAND, name="Arial")
SUB_FONT = Font(italic=True, color="555555", name="Arial")
KPI_LBL_FONT = Font(bold=True, size=10, color="FFFFFF", name="Arial")
KPI_VAL_FONT = Font(bold=True, size=22, color=BRAND, name="Arial")
DEFAULT_FONT = Font(name="Arial", size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header_row(ws, row, last_col):
    for c in range(1, last_col+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER

def auto_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# =================================================================
# README sheet
# =================================================================
ws = wb.active
ws.title = "README"
ws.sheet_view.showGridLines = False
ws["A1"] = "Project Portfolio Tracker — Excel Dashboard"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Companion to the Excel + SQL portfolio project. Built on synthetic data."
ws["A2"].font = SUB_FONT
ws["A4"] = "Tabs in this workbook"
ws["A4"].font = Font(bold=True, name="Arial", size=12)

readme_rows = [
    ("Dashboard",          "Executive view: KPI cards, project health chart, budget burn, capacity."),
    ("Exec_OnePager",      "Print-ready 1-page summary you can attach to LinkedIn or email."),
    ("Data_Projects",      "Cleaned project list with computed % complete, % budget burned, health flag."),
    ("Data_Tasks",         "Task-level data with overdue flag."),
    ("Data_TimeEntries_Wk","Weekly hours per employee with capacity flag."),
    ("Data_Budget",        "Monthly budget vs. actual per project with variance."),
]
ws["A6"] = "Tab"; ws["B6"] = "What's there"
style_header_row(ws, 6, 2)
for i, (tab, desc) in enumerate(readme_rows, start=7):
    ws.cell(row=i, column=1, value=tab).font = Font(bold=True, name="Arial")
    ws.cell(row=i, column=2, value=desc).font = DEFAULT_FONT

ws["A14"] = "How to refresh"
ws["A14"].font = Font(bold=True, name="Arial", size=12)
ws["A15"] = "1. Replace data in the Data_* tabs (paste from SQL output)."
ws["A16"] = "2. All formulas, KPIs, and charts on the Dashboard recalculate automatically."
ws["A17"] = "3. Update the 'As of' date in cell B3 of the Dashboard tab."
for r in range(15,18): ws.cell(row=r, column=1).font = DEFAULT_FONT

auto_widths(ws, {"A":24,"B":80})

# =================================================================
# Data_Projects
# =================================================================
ws = wb.create_sheet("Data_Projects")
headers = ["project_id","project_name","client_name","project_manager",
           "start_date","planned_end_date","actual_end_date","status",
           "budget_amount","actual_amount",
           "pct_complete","pct_budget_burned","health_flag","overrun"]
ws.append(headers)
style_header_row(ws, 1, len(headers))

for p in projects:
    pid = int(p["project_id"])
    pm_id = p["project_manager_id"]
    pm = emp_by_id.get(int(pm_id), {}).get("full_name") if pm_id else "(unassigned)"
    bud = budget_by_proj[pid]
    row = [
        pid,
        p["project_name"],
        p["client_name"],
        pm,
        parse_date(p["start_date"]),
        parse_date(p["planned_end_date"]),
        parse_date(p["actual_end_date"]) if p["actual_end_date"] else None,
        clean_status(p["status"]),
        float(p["budget_amount"]),
        round(bud["actual"], 2),
        None,  # pct_complete (formula below)
        None,  # pct_budget_burned (formula below)
        None,  # health_flag (formula below)
        None,  # overrun (formula below)
    ]
    ws.append(row)

proj_count = len(projects)
for r in range(2, 2 + proj_count):
    ws.cell(row=r, column=11, value=(
        f'=IFERROR(COUNTIFS(Data_Tasks!B:B,A{r},Data_Tasks!G:G,"Done")'
        f'/COUNTIFS(Data_Tasks!B:B,A{r}),0)'
    ))
    ws.cell(row=r, column=12, value=f"=IFERROR(J{r}/I{r},0)")
    ws.cell(row=r, column=13, value=(
        f'=IF(J{r}>I{r},"Over Budget",'
        f'IF((L{r}-K{r})>0.15,"At Risk","On Track"))'
    ))
    ws.cell(row=r, column=14, value=f"=J{r}-I{r}")

for r in range(2, 2 + proj_count):
    ws.cell(row=r, column=9).number_format = '"$"#,##0;[Red]("$"#,##0);"-"'
    ws.cell(row=r, column=10).number_format = '"$"#,##0;[Red]("$"#,##0);"-"'
    ws.cell(row=r, column=11).number_format = "0.0%"
    ws.cell(row=r, column=12).number_format = "0.0%"
    ws.cell(row=r, column=14).number_format = '"$"#,##0;[Red]("$"#,##0);"-"'
    for c in (5,6,7):
        ws.cell(row=r, column=c).number_format = "yyyy-mm-dd"

last_row = 1 + proj_count
green = PatternFill("solid", fgColor="C6EFCE")
yellow = PatternFill("solid", fgColor="FFEB9C")
red = PatternFill("solid", fgColor="F8CBAD")
ws.conditional_formatting.add(
    f"M2:M{last_row}", CellIsRule(operator="equal", formula=['"On Track"'], fill=green))
ws.conditional_formatting.add(
    f"M2:M{last_row}", CellIsRule(operator="equal", formula=['"At Risk"'], fill=yellow))
ws.conditional_formatting.add(
    f"M2:M{last_row}", CellIsRule(operator="equal", formula=['"Over Budget"'], fill=red))

tbl = Table(displayName="tProjects", ref=f"A1:N{last_row}")
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(tbl)

auto_widths(ws, {"A":11,"B":28,"C":20,"D":22,"E":12,"F":13,"G":13,
                 "H":13,"I":13,"J":13,"K":12,"L":15,"M":13,"N":13})
ws.freeze_panes = "A2"

# =================================================================
# Data_Tasks
# =================================================================
ws = wb.create_sheet("Data_Tasks")
headers = ["task_id","project_id","task_name","assignee","due_date","completed_date",
           "status","estimated_hours","overdue_flag"]
ws.append(headers)
style_header_row(ws, 1, len(headers))

for t in tasks:
    asn_id = t["assignee_id"]
    asn = emp_by_id.get(int(asn_id), {}).get("full_name") if asn_id else ""
    ws.append([
        int(t["task_id"]),
        int(t["project_id"]),
        t["task_name"],
        asn,
        parse_date(t["due_date"]),
        parse_date(t["completed_date"]) if t["completed_date"] else None,
        t["status"],
        float(t["estimated_hours"]),
        None,
    ])

n = len(tasks)
for r in range(2, 2+n):
    ws.cell(row=r, column=9, value=f'=IF(AND(G{r}<>"Done",E{r}<TODAY()),"Overdue","")')
    for c in (5,6):
        ws.cell(row=r, column=c).number_format = "yyyy-mm-dd"

ws.conditional_formatting.add(
    f"I2:I{1+n}", CellIsRule(operator="equal", formula=['"Overdue"'], fill=red))

tbl = Table(displayName="tTasks", ref=f"A1:I{1+n}")
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(tbl)
auto_widths(ws, {"A":9,"B":11,"C":34,"D":22,"E":12,"F":14,"G":13,"H":12,"I":11})
ws.freeze_panes = "A2"

# =================================================================
# Data_TimeEntries_Wk
# =================================================================
ws = wb.create_sheet("Data_TimeEntries_Wk")
headers = ["employee_id","employee","department","week_start","hours_this_week","capacity_flag"]
ws.append(headers)
style_header_row(ws, 1, len(headers))

rows_sorted = sorted(hours_wk.items(), key=lambda kv: (kv[0][1], kv[0][0]))
for (eid, wk), hrs in rows_sorted:
    e = emp_by_id[eid]
    ws.append([eid, e["full_name"], e["department"], wk, round(hrs,2), None])

n = len(rows_sorted)
for r in range(2, 2+n):
    ws.cell(row=r, column=4).number_format = "yyyy-mm-dd"
    ws.cell(row=r, column=6, value=(
        f'=IF(E{r}>45,"Overallocated",IF(E{r}<30,"Underutilized","OK"))'
    ))

ws.conditional_formatting.add(
    f"F2:F{1+n}", CellIsRule(operator="equal", formula=['"Overallocated"'], fill=red))
ws.conditional_formatting.add(
    f"F2:F{1+n}", CellIsRule(operator="equal", formula=['"Underutilized"'], fill=yellow))
ws.conditional_formatting.add(
    f"F2:F{1+n}", CellIsRule(operator="equal", formula=['"OK"'], fill=green))

tbl = Table(displayName="tTime", ref=f"A1:F{1+n}")
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(tbl)
auto_widths(ws, {"A":11,"B":22,"C":14,"D":12,"E":15,"F":15})
ws.freeze_panes = "A2"

# =================================================================
# Data_Budget
# =================================================================
ws = wb.create_sheet("Data_Budget")
headers = ["record_id","project_id","project_name","month_start",
           "budgeted_amount","actual_amount","variance"]
ws.append(headers)
style_header_row(ws, 1, len(headers))

proj_name_by_id = {int(p["project_id"]): p["project_name"] for p in projects}
for b in budget:
    pid = int(b["project_id"])
    ws.append([
        int(b["record_id"]),
        pid,
        proj_name_by_id.get(pid, ""),
        parse_date(b["month_start"]),
        float(b["budgeted_amount"]),
        float(b["actual_amount"]),
        None,
    ])

n = len(budget)
for r in range(2, 2+n):
    ws.cell(row=r, column=4).number_format = "yyyy-mm-dd"
    ws.cell(row=r, column=5).number_format = '"$"#,##0;[Red]("$"#,##0);"-"'
    ws.cell(row=r, column=6).number_format = '"$"#,##0;[Red]("$"#,##0);"-"'
    ws.cell(row=r, column=7, value=f"=F{r}-E{r}").number_format = '"$"#,##0;[Red]("$"#,##0);"-"'

ws.conditional_formatting.add(
    f"G2:G{1+n}",
    ColorScaleRule(start_type="min", start_color="63BE7B",
                   mid_type="num", mid_value=0, mid_color="FFFFFF",
                   end_type="max", end_color="F8696B"))

tbl = Table(displayName="tBudget", ref=f"A1:G{1+n}")
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(tbl)
auto_widths(ws, {"A":11,"B":11,"C":28,"D":12,"E":15,"F":15,"G":15})
ws.freeze_panes = "A2"

# =================================================================
# Dashboard
# =================================================================
ws = wb.create_sheet("Dashboard")
ws.sheet_view.showGridLines = False

ws["B2"] = "Project Portfolio — Health Dashboard"
ws["B2"].font = TITLE_FONT
ws["B3"] = "As of:"
ws["B3"].font = Font(bold=True, name="Arial")
ws["C3"] = date(2026,5,2)
ws["C3"].number_format = "yyyy-mm-dd"
ws["C3"].font = Font(name="Arial")
ws["B4"] = "Synthetic data | Built with Excel + SQL"
ws["B4"].font = SUB_FONT

def kpi_card(ws, row, col, label, value_formula, fill_color=BRAND, val_format="0"):
    label_cell = ws.cell(row=row, column=col, value=label)
    label_cell.fill = PatternFill("solid", fgColor=fill_color)
    label_cell.font = KPI_LBL_FONT
    label_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
    val_cell = ws.cell(row=row+1, column=col, value=value_formula)
    val_cell.font = KPI_VAL_FONT
    val_cell.alignment = Alignment(horizontal="center", vertical="center")
    val_cell.number_format = val_format
    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+2, end_column=col+1)
    val_cell.border = BORDER
    label_cell.border = BORDER

for col in range(2, 12):
    ws.column_dimensions[get_column_letter(col)].width = 13
ws.row_dimensions[6].height = 22
ws.row_dimensions[7].height = 32
ws.row_dimensions[8].height = 12

kpi_card(ws, 6, 2, "TOTAL PROJECTS", "=COUNTA(tProjects[project_id])")
kpi_card(ws, 6, 4, "ON TRACK",
         '=COUNTIFS(tProjects[health_flag],"On Track")', fill_color="2E7D32")
kpi_card(ws, 6, 6, "AT RISK",
         '=COUNTIFS(tProjects[health_flag],"At Risk")', fill_color="B26A00")
kpi_card(ws, 6, 8, "OVER BUDGET",
         '=COUNTIFS(tProjects[health_flag],"Over Budget")', fill_color="B71C1C")
kpi_card(ws, 6, 10, "OVERDUE TASKS",
         '=COUNTIFS(tTasks[overdue_flag],"Overdue")', fill_color="6A1B9A")

ws["B11"] = "Project Health Summary"
ws["B11"].font = Font(bold=True, size=12, name="Arial", color=BRAND)

ws["B12"] = "Health"; ws["C12"] = "# Projects"; ws["D12"] = "% of Portfolio"
style_header_row(ws, 12, 4)

flags = ["On Track", "At Risk", "Over Budget"]
for i, fl in enumerate(flags):
    r = 13 + i
    ws.cell(row=r, column=2, value=fl)
    ws.cell(row=r, column=3, value=f'=COUNTIFS(tProjects[health_flag],B{r})')
    ws.cell(row=r, column=4, value=f'=IFERROR(C{r}/COUNTA(tProjects[project_id]),0)')
    ws.cell(row=r, column=4).number_format = "0.0%"

ws.conditional_formatting.add(
    "B13:B15", CellIsRule(operator="equal", formula=['"On Track"'], fill=green))
ws.conditional_formatting.add(
    "B13:B15", CellIsRule(operator="equal", formula=['"At Risk"'], fill=yellow))
ws.conditional_formatting.add(
    "B13:B15", CellIsRule(operator="equal", formula=['"Over Budget"'], fill=red))

chart = BarChart()
chart.type = "bar"
chart.style = 11
chart.title = "Projects by Health Flag"
chart.y_axis.title = "Health"
chart.x_axis.title = "# Projects"
data_ref = Reference(ws, min_col=3, min_row=12, max_row=15, max_col=3)
cat_ref = Reference(ws, min_col=2, min_row=13, max_row=15)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cat_ref)
chart.height = 7
chart.width = 14
chart.dataLabels = DataLabelList(showVal=True)
ws.add_chart(chart, "F11")

ws["B19"] = "Top 5 Over-Budget Projects"
ws["B19"].font = Font(bold=True, size=12, name="Arial", color=BRAND)
ws["B20"] = "Project"; ws["C20"] = "Client"; ws["D20"] = "Budget"; ws["E20"] = "Actual"; ws["F20"] = "Overrun"
style_header_row(ws, 20, 5)

for i in range(5):
    r = 21 + i
    rank = i + 1
    rng_overrun = "tProjects[overrun]"
    rng_name    = "tProjects[project_name]"
    rng_client  = "tProjects[client_name]"
    rng_budget  = "tProjects[budget_amount]"
    rng_actual  = "tProjects[actual_amount]"
    ws.cell(row=r, column=2, value=(
        f'=IFERROR(IF(LARGE({rng_overrun},{rank})<=0,"",'
        f'INDEX({rng_name},MATCH(LARGE({rng_overrun},{rank}),{rng_overrun},0))),"")'
    ))
    ws.cell(row=r, column=3, value=(
        f'=IFERROR(IF(LARGE({rng_overrun},{rank})<=0,"",'
        f'INDEX({rng_client},MATCH(LARGE({rng_overrun},{rank}),{rng_overrun},0))),"")'
    ))
    ws.cell(row=r, column=4, value=(
        f'=IFERROR(IF(LARGE({rng_overrun},{rank})<=0,"",'
        f'INDEX({rng_budget},MATCH(LARGE({rng_overrun},{rank}),{rng_overrun},0))),"")'
    ))
    ws.cell(row=r, column=5, value=(
        f'=IFERROR(IF(LARGE({rng_overrun},{rank})<=0,"",'
        f'INDEX({rng_actual},MATCH(LARGE({rng_overrun},{rank}),{rng_overrun},0))),"")'
    ))
    ws.cell(row=r, column=6, value=(
        f'=IFERROR(IF(LARGE({rng_overrun},{rank})<=0,"",LARGE({rng_overrun},{rank})),"")'
    ))
    for c in (4,5,6):
        ws.cell(row=r, column=c).number_format = '"$"#,##0;[Red]("$"#,##0);"-"'

ws["B28"] = "Monthly Spend — Planned vs Actual"
ws["B28"].font = Font(bold=True, size=12, name="Arial", color=BRAND)

months_sorted = sorted({parse_date(b["month_start"]) for b in budget})
ws["I20"] = "Month"; ws["J20"] = "Planned"; ws["K20"] = "Actual"
for c in range(9,12):
    ws.cell(row=20, column=c).fill = HDR_FILL
    ws.cell(row=20, column=c).font = HDR_FONT

for i, m in enumerate(months_sorted):
    r = 21 + i
    ws.cell(row=r, column=9, value=m).number_format = "mmm yyyy"
    ws.cell(row=r, column=10, value=f'=SUMIFS(tBudget[budgeted_amount],tBudget[month_start],I{r})')
    ws.cell(row=r, column=11, value=f'=SUMIFS(tBudget[actual_amount],tBudget[month_start],I{r})')
    ws.cell(row=r, column=10).number_format = '"$"#,##0'
    ws.cell(row=r, column=11).number_format = '"$"#,##0'

last_month_row = 20 + len(months_sorted)
line = LineChart()
line.title = "Planned vs Actual by Month"
line.y_axis.title = "$"
line.x_axis.title = "Month"
line.height = 7
line.width = 14
data_ref = Reference(ws, min_col=10, min_row=20, max_row=last_month_row, max_col=11)
cat_ref = Reference(ws, min_col=9, min_row=21, max_row=last_month_row)
line.add_data(data_ref, titles_from_data=True)
line.set_categories(cat_ref)
ws.add_chart(line, "B30")

for c in ("I","J","K"): ws.column_dimensions[c].width = 13

ws["B46"] = "Capacity Hot List — Overallocated Employees (any week >45 hrs)"
ws["B46"].font = Font(bold=True, size=12, name="Arial", color=BRAND)
ws["B47"] = "Employee"; ws["C47"] = "Department"; ws["D47"] = "# Weeks Overallocated"
style_header_row(ws, 47, 4)

unique_emp_ids = sorted({e["employee_id"] for e in employees}, key=int)
helper_start = 60
ws[f"B{helper_start}"] = "(helper)"
ws[f"B{helper_start}"].font = SUB_FONT
ws[f"C{helper_start}"] = "employee_id"
ws[f"D{helper_start}"] = "name"
ws[f"E{helper_start}"] = "dept"
ws[f"F{helper_start}"] = "overalloc_wks"
for i, eid in enumerate(unique_emp_ids):
    r = helper_start + 1 + i
    e = emp_by_id[int(eid)]
    ws.cell(row=r, column=3, value=int(eid))
    ws.cell(row=r, column=4, value=e["full_name"])
    ws.cell(row=r, column=5, value=e["department"])
    ws.cell(row=r, column=6, value=(
        f'=COUNTIFS(tTime[employee_id],C{r},tTime[capacity_flag],"Overallocated")'
        f'+C{r}/1000000'
    ))
helper_end = helper_start + len(unique_emp_ids)

for i in range(10):
    r = 48 + i
    rank = i + 1
    rng_count = f"F{helper_start+1}:F{helper_end}"
    rng_name = f"D{helper_start+1}:D{helper_end}"
    rng_dept = f"E{helper_start+1}:E{helper_end}"
    ws.cell(row=r, column=2, value=(
        f'=IFERROR(IF(INT(LARGE({rng_count},{rank}))=0,"",'
        f'INDEX({rng_name},MATCH(LARGE({rng_count},{rank}),{rng_count},0))),"")'
    ))
    ws.cell(row=r, column=3, value=(
        f'=IFERROR(IF(INT(LARGE({rng_count},{rank}))=0,"",'
        f'INDEX({rng_dept},MATCH(LARGE({rng_count},{rank}),{rng_count},0))),"")'
    ))
    ws.cell(row=r, column=4, value=(
        f'=IFERROR(IF(INT(LARGE({rng_count},{rank}))=0,"",INT(LARGE({rng_count},{rank}))),"")'
    ))

ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 15
ws.column_dimensions["D"].width = 22

for r in range(helper_start, helper_end+1):
    for c in range(3,7):
        ws.cell(row=r, column=c).font = Font(color="999999", size=9, name="Arial")

# =================================================================
# Exec_OnePager
# =================================================================
ws = wb.create_sheet("Exec_OnePager")
ws.sheet_view.showGridLines = False

ws["B2"] = "Project Portfolio Health — Executive Summary"
ws["B2"].font = TITLE_FONT
ws["B3"] = "Synthetic dataset | Sample portfolio piece"
ws["B3"].font = SUB_FONT

for col in range(2, 12): ws.column_dimensions[get_column_letter(col)].width = 13
ws.row_dimensions[5].height = 22
ws.row_dimensions[6].height = 32
ws.row_dimensions[7].height = 12

kpi_card(ws, 5, 2, "PROJECTS", "=COUNTA(tProjects[project_id])")
kpi_card(ws, 5, 4, "% ON TRACK",
         '=IFERROR(COUNTIFS(tProjects[health_flag],"On Track")/COUNTA(tProjects[project_id]),0)',
         fill_color="2E7D32", val_format="0.0%")
kpi_card(ws, 5, 6, "% OVER BUDGET",
         '=IFERROR(COUNTIFS(tProjects[health_flag],"Over Budget")/COUNTA(tProjects[project_id]),0)',
         fill_color="B71C1C", val_format="0.0%")
kpi_card(ws, 5, 8, "TOTAL OVERRUN",
         '=SUMIFS(tProjects[overrun],tProjects[overrun],">0")',
         fill_color="6A1B9A", val_format='"$"#,##0')

ws["B10"] = "What this dashboard shows"
ws["B10"].font = Font(bold=True, size=12, name="Arial", color=BRAND)
narrative = [
    "• Single weekly view of project health across all 20 projects.",
    "• Health flag is automatic: 'Over Budget' when actuals exceed plan; 'At Risk' when",
    "  budget burn is running >15 percentage points ahead of % complete.",
    "• Capacity hot list flags employees logging >45 hours for any week in the last 8.",
]
for i, line_t in enumerate(narrative):
    ws.cell(row=11+i, column=2, value=line_t).font = DEFAULT_FONT
    ws.merge_cells(start_row=11+i, start_column=2, end_row=11+i, end_column=11)

ws["B16"] = "Recommended actions"
ws["B16"].font = Font(bold=True, size=12, name="Arial", color=BRAND)
actions = [
    "1. Re-baseline the over-budget projects with the COO this week.",
    "2. Open weekly check-in on any project where (% burned − % complete) > 15.",
    "3. Redistribute work from overallocated staff before quality slips.",
    "4. Replace ad-hoc PM emails with this dashboard, refreshed every Monday.",
]
for i, line_t in enumerate(actions):
    ws.cell(row=17+i, column=2, value=line_t).font = DEFAULT_FONT
    ws.merge_cells(start_row=17+i, start_column=2, end_row=17+i, end_column=11)

ws["B23"] = "Built by Brystal | Excel + SQL portfolio piece"
ws["B23"].font = SUB_FONT

ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.print_options.horizontalCentered = True
ws.page_margins.left = 0.5
ws.page_margins.right = 0.5

# =================================================================
# Default font on every sheet
# =================================================================
for sheet in wb.sheetnames:
    s = wb[sheet]
    for row in s.iter_rows():
        for cell in row:
            if cell.font and cell.font.name in (None, "Calibri"):
                cell.font = Font(
                    name="Arial",
                    size=cell.font.size or 10,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    color=cell.font.color,
                )

# =================================================================
# Save
# =================================================================
OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")
print(f"Tabs : {wb.sheetnames}")
